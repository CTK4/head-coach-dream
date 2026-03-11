#!/usr/bin/env python3
import argparse
import collections
import datetime
import json
import math
from pathlib import Path

TITLE_SHEETS = {'PlayerMorale','QBWRChemistry','TeamLeadershipImpact','MoraleConfig','MoraleFormulas'}
DRAFT_HEADER_MAP = {'Player ID':'playerId', 40.0:'Forty'}
TECH_FIELDS = [
    'Accuracy','Arm_Strength','Touch','Mechanics','Pocket_Presence','Release',
    'Vision','Balance','Elusiveness','Hands','Pass_Protection','Footwork',
    'Route_Running','YAC_Ability','Body_Control','Contested_Catch',
    'Blocking','Pass_Blocking','Run_Blocking','Hand_Technique','Anchor','Mirror_Ability',
    'Power','Lateral_Movement','Pass_Rush','Run_Defense','Leverage','Gap_Discipline','Motor2',
    'Bend','Speed_to_Power','Pursuit','Tackling','Coverage','Blitz_Ability','Shed_Blocks',
    'Man_Coverage','Zone_Coverage','Press_Technique','Ball_Skills','Backpedal','Run_Support',
    'Range','Angles','Kick_Accuracy','Kick_Power'
]
MENTAL_COEFFS = {
    'Awareness': 0.3024,
    'Poise': 0.1401,
    'Focus': 0.0961,
    'Leadership': 0.0033,
    'Competitiveness': 0.0034,
    'Confidence': 0.0114,
    'Work_Ethic': 0.0102,
    'Coachability': -0.0193,
    'Football_IQ': 0.2464,
    'Instincts': 0.2038,
    'Maturity': 0.0005,
    'Motor': -0.0009,
}
MENTAL_INTERCEPT = 0.1711
ATHLETIC_COEFFS = {
    'Speed': 0.2011,
    'Acceleration': 0.3790,
    'Agility': 0.1645,
    'Strength': 0.0064,
    'Jumping': 0.0865,
    'Stamina': 0.1589,
}
ATHLETIC_INTERCEPT = 0.1730

def maybe_convert_datetime(key, value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        if key == 'scheme':
            return '4-2-5'
        if isinstance(value, datetime.datetime):
            return value.strftime('%Y-%m-%d')
        return value.isoformat()
    return value

def clean_number(value):
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        if value.is_integer():
            return int(value)
        return float(value)
    return value

def normalize_value(key, value):
    value = maybe_convert_datetime(key, value)
    return clean_number(value)

def parse_workbook(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)
        title = None
        if sheet in TITLE_SHEETS:
            first = next(rows_iter)
            title = first[0] if first else None
        raw_headers = next(rows_iter)
        headers = [DRAFT_HEADER_MAP.get(h, h) if sheet == '2026_Draft_Class' else h for h in raw_headers]
        rows = []
        for row in rows_iter:
            if not row or all(v is None for v in row):
                continue
            item = {}
            for h, v in zip(headers, row):
                if h is None:
                    continue
                item[h] = normalize_value(h, v)
            rows.append(item)
        out[sheet] = (title, headers, rows)
    return out

def derive_divisions_from_teams(teams):
    seen = collections.OrderedDict()
    sort_orders = collections.defaultdict(int)
    for team in teams:
        div_id = team['divisionId']
        conf_id = team['conferenceId']
        if div_id not in seen:
            sort_orders[conf_id] += 1
            name = div_id.split('_', 1)[1].replace('_', ' ').title()
            seen[div_id] = {
                'divisionId': div_id,
                'conferenceId': conf_id,
                'name': name,
                'sortOrder': sort_orders[conf_id],
                'notes': None,
            }
    return list(seen.values())

def compute_mental(player):
    total = MENTAL_INTERCEPT
    for field, coeff in MENTAL_COEFFS.items():
        v = player.get(field)
        if v is None:
            return None
        total += coeff * v
    return round(total, 2)

def compute_athleticism(player):
    total = ATHLETIC_INTERCEPT
    for field, coeff in ATHLETIC_COEFFS.items():
        v = player.get(field)
        if v is None:
            return None
        total += coeff * v
    return round(total, 2)

def compute_tech_core(player):
    vals = [player.get(field) for field in TECH_FIELDS if player.get(field) is not None]
    if not vals:
        return None
    return round(sum(vals) / len(vals), 8)

def draft_tier_from_rank(rank):
    if rank is None:
        return None
    if rank <= 2:
        return 'Top 5'
    if rank <= 20:
        return 'Round1'
    if rank <= 30:
        return 'Round2'
    if rank <= 40:
        return 'Round3'
    if rank <= 50:
        return 'Round4'
    if rank <= 60:
        return 'Round5'
    if rank <= 70:
        return 'Round6'
    if rank <= 80:
        return 'Round7'
    return 'UDFA'

def synthesize_player_morale(players, existing_rows, season):
    by_id = {row['playerId']: row for row in existing_rows}
    for p in players:
        pid = p['playerId']
        if pid in by_id:
            continue
        by_id[pid] = {
            'playerId': pid,
            'fullName': p.get('fullName'),
            'pos': p.get('pos'),
            'teamId': p.get('teamId'),
            'season': season,
            'week': 0,
            'morale': p.get('morale'),
            'morale_trend': p.get('morale_trend'),
            'role_expectation': p.get('role_expectation'),
            'playing_time_satisfaction': p.get('playing_time_satisfaction'),
            'contract_satisfaction': p.get('contract_satisfaction'),
            'trade_request_state': p.get('trade_request_state'),
            'discipline_risk': p.get('discipline_risk'),
            'locker_room_toxicity_risk': p.get('locker_room_toxicity_risk'),
            'LIS': p.get('LIS'),
            'player_usage_share': 0,
            'wins_last4': 0,
            'is_injured': False,
            'had_discipline_event': False,
            'is_team_captain': p.get('Leadership_Role') == 'Team Captain',
            'delta_morale_last': 0,
            'TRP': 0,
            'consecutive_high_trp_weeks': 0,
        }
    ordered = []
    seen = set()
    for row in existing_rows:
        ordered.append(row)
        seen.add(row['playerId'])
    for p in players:
        if p['playerId'] not in seen:
            ordered.append(by_id[p['playerId']])
            seen.add(p['playerId'])
    return ordered

def legacy_title_sheet_export(clean_title, clean_headers, clean_rows):
    legacy_headers = [clean_title] + [f'Unnamed: {i}' for i in range(1, len(clean_headers))]
    legacy_rows = []
    legacy_rows.append({k: clean_number(v) for k, v in zip(legacy_headers, clean_headers)})
    for row in clean_rows:
        values = [row.get(h) for h in clean_headers]
        legacy_rows.append({k: clean_number(v) for k, v in zip(legacy_headers, values)})
    return legacy_rows

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx', required=True)
    parser.add_argument('--current-json', required=True)
    parser.add_argument('--out', required=True)
    parser.add_argument('--legacy-title-sheets', action='store_true')
    args = parser.parse_args()

    with open(args.current_json) as f:
        current = json.load(f)

    parsed = parse_workbook(args.xlsx)
    top_order = list(current.keys())
    player_key_order = list(current['Players'][0].keys())
    personnel_key_order = list(current['Personnel'][0].keys()) + ['Column1']
    draft_key_order = list(current['2026_Draft_Class'][0].keys())

    teams = parsed['Teams'][2]
    divisions = derive_divisions_from_teams(teams)

    players = []
    for p in parsed['Players'][2]:
        merged = dict(p)
        merged['Mental_Core'] = compute_mental(merged)
        merged['Athleticism'] = compute_athleticism(merged)
        merged['Tech_Core'] = compute_tech_core(merged)
        ordered = {}
        for key in player_key_order:
            ordered[key] = merged.get(key)
        for key, value in merged.items():
            if key not in ordered:
                ordered[key] = value
        players.append(ordered)

    season = parsed['League'][2][0]['season']
    player_morale = synthesize_player_morale(players, parsed['PlayerMorale'][2], season)

    personnel = []
    for row in parsed['Personnel'][2]:
        ordered = {}
        for key in personnel_key_order:
            if key in row:
                ordered[key] = row.get(key)
        for k, v in row.items():
            if k not in ordered:
                ordered[k] = v
        personnel.append(ordered)

    draft_class = []
    for row in parsed['2026_Draft_Class'][2]:
        merged = dict(row)
        merged['DraftTier'] = draft_tier_from_rank(merged.get('Rank'))
        ordered = {}
        for key in draft_key_order:
            ordered[key] = merged.get(key)
        if 'overall' in merged:
            ordered['overall'] = merged['overall']
        for key, value in merged.items():
            if key not in ordered:
                ordered[key] = value
        draft_class.append(ordered)

    export = {
        'League': parsed['League'][2],
        'Conferences': parsed['Conferences'][2],
        'Divisions': divisions,
        'Teams': teams,
        'Personnel': personnel,
        'Players': players,
        'PlayerMorale': player_morale,
        'QBWRChemistry': parsed['QBWRChemistry'][2],
        'TeamLeadershipImpact': parsed['TeamLeadershipImpact'][2],
        'MoraleConfig': parsed['MoraleConfig'][2],
        'MoraleFormulas': parsed['MoraleFormulas'][2],
        'Contracts': parsed['Contracts'][2],
        'DraftOrder': parsed['DraftOrder'][2],
        'DraftPicks': parsed['DraftPicks'][2],
        'TeamFinances': parsed['TeamFinances'][2],
        '2026_Draft_Class': draft_class,
        'Combine_Metrics': parsed['Combine_Metrics'][2],
    }
    export = {k: export[k] for k in top_order}

    if args.legacy_title_sheets:
        for sheet in TITLE_SHEETS:
            title, headers, rows = parsed[sheet]
            if sheet == 'PlayerMorale':
                rows = player_morale
            export[sheet] = legacy_title_sheet_export(title, headers, rows)

    with open(args.out, 'w') as f:
        json.dump(export, f, indent=2, ensure_ascii=False)

if __name__ == '__main__':
    main()
